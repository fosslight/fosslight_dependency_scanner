#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Copyright (c) 2020 LG Electronics Inc.
# SPDX-License-Identifier: Apache-2.0

from __future__ import print_function
from io import open
import os
import sys
import argparse
from openpyxl import load_workbook, Workbook
import platform
import shutil
import subprocess
import json
import re
from xml.etree.ElementTree import parse
from bs4 import BeautifulSoup
import logging
import requests

VERSION = "3.0.0"

# Check the manifest file
manifest_array = [["pip", "requirements.txt"], ["npm", "package.json"], ["maven", "pom.xml"],
                  ["gradle", "build.gradle"], ["pub", "pubspec.yaml"]]

# binary url to check license text
license_scanner_url_linux = "https://github.com/LGE-OSS/fosslight_dependency/raw/main/third_party/nomos/nomossa"
license_scanner_url_macos = "https://github.com/LGE-OSS/fosslight_dependency/raw/main/third_party/askalono/askalono_macos"
license_scanner_url_windows = "https://github.com/LGE-OSS/fosslight_dependency/raw/main/third_party/askalono/askalono.exe"


class HelpStop(Exception):
    pass


def check_valid_manifest_file():
    global PACKAGE

    manifest_file_name = [i[1] for i in manifest_array]

    idx = 0
    found_idx = []
    for f in manifest_file_name:
        if os.path.isfile(f):
            found_idx.append(idx)
        idx += 1

    if len(found_idx) == 1:
        PACKAGE = manifest_array[int(found_idx[0])][0]
        logging.info("### Info Message ###")
        logging.info("Found the manifest file(" + manifest_array[int(found_idx[0])][1] + ")automatically.")
        logging.info("Set PACKAGE =" + PACKAGE)
        ret = 0
    else:
        ret = 1

    return ret


def help_print():
    logging.info("### Option Usage ###")
    logging.info(" -h : print usage message")
    logging.info(" -v : print the version of the script")
    logging.info(" -m : enter the package manager")
    logging.info("      ex) pip, npm, maven, gradle, pub")
    logging.info(" -c : enter the customized build output directory of maven, gradle")
    logging.info("      ** The default build output directory of maven is 'target', and which of gradle is 'build'.")
    logging.info("        If you use the customized build output directory, then use this option with your output directory name.")
    logging.info(" -p : enter the path where the script will be run.")
    logging.info(" -o : enter the path where the result file will be generated.")


def parse_option():
    global MANUAL_DETECT, PIP_ACTIVATE, PIP_DEACTIVATE, PACKAGE, OUTPUT_CUSTOM_DIR, CUR_PATH, OUTPUT_RESULT_DIR

    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument('-h', '--help', action='store_true', required=False)
    parser.add_argument('-m', '--manager', nargs=1, type=str, default="UNSPECIFIED", required=False)
    parser.add_argument('-a', '--activate', nargs=1, type=str, default="UNSPECIFIED", required=False)
    parser.add_argument('-d', '--deactivate', nargs=1, type=str, default="UNSPECIFIED", required=False)
    parser.add_argument('-c', '--customized', nargs=1, type=str, required=False)
    parser.add_argument('-p', '--path', nargs=1, type=str, required=False)
    parser.add_argument('-v', '--version', action='store_true', required=False)
    parser.add_argument('-o', '--output', nargs=1, type=str, required=False)

    args = parser.parse_args()

    # -h option
    if args.help:
        help_print()
        raise HelpStop

    # -v option
    if args.version:
        logging.info(VERSION)
        raise HelpStop

    # -m option
    if args.manager == "UNSPECIFIED":
        MANUAL_DETECT = 0
        # It will be detected the package manager automatically with manifest file.
    else:
        MANUAL_DETECT = 1
        PACKAGE = "".join(args.manager)

    # -a option
    if args.activate:
        PIP_ACTIVATE = "".join(args.activate)

    # -d option
    if args.deactivate:
        PIP_DEACTIVATE = "".join(args.deactivate)

    # -c option
    if args.customized:
        OUTPUT_CUSTOM_DIR = "".join(args.customized)
    else:
        OUTPUT_CUSTOM_DIR = ""

    # -o option
    if args.output:
        OUTPUT_RESULT_DIR = "".join(args.output)
        if os.path.isdir(OUTPUT_RESULT_DIR):
            OUTPUT_RESULT_DIR = os.path.abspath(OUTPUT_RESULT_DIR)
        else:
            try:
                os.mkdir(OUTPUT_RESULT_DIR)
            except:
                logging.error("You entered wrong output path(" + OUTPUT_RESULT_DIR + ") to generate output file.")
                sys.exit(1)
            OUTPUT_RESULT_DIR = os.path.abspath(OUTPUT_RESULT_DIR)
    else:
        OUTPUT_RESULT_DIR = os.getcwd()

    # -p option
    if args.path:
        CUR_PATH = "".join(args.path)
        if os.path.isdir(CUR_PATH):
            os.chdir(CUR_PATH)
            CUR_PATH = os.getcwd()
        else:
            logging.error("You entered wrong path(" + CUR_PATH + ") to run the script.")
            sys.exit(1)
    else:
        CUR_PATH = os.getcwd()
        os.chdir(CUR_PATH)

    logging.info("This script is from the path(" + CUR_PATH + ")")


def configure_package():
    parse_option()
    if MANUAL_DETECT == 0:
        ret = check_valid_manifest_file()
        if ret != 0:
            logging.error("### Error Message ###")
            logging.error("Please enter the package manager with -m option.")
            logging.error("Choose your package manager. [npm, pip, maven, gradle]")
            logging.error(">>Command Example")
            logging.error("  python dependency_unified.py -m npm")
            sys.exit(1)


# Common variables for OSS report
OSS_report_1st_row = ['ID', 'Source Name or Path', 'OSS Name', 'OSS Version', 'License', 'Download Location',
                      'Homepage', 'Copyright Text', 'License Text', 'Exclude', 'Comment']
OSS_report_2nd_row = ['-', '[Name of the Source File or Path]', '[Name of the OSS used in the Source Code]',
                      '[Version Number of the OSS]',
                      '[License of the OSS. Use SPDX Identifier : https://spdx.org/licenses/]',
                      '[Download URL or a specific location within a VCS for the OSS]',
                      '[Web site that serves as the OSS\'s home page]', '[The copyright holders of the OSS]',
                      '[License Text of the License. This field can be skipped if the License is in SPDX.]',
                      '[If this OSS is not included in the final version, Exclude]']


####################
# Common functions #
####################

def check_python_version():
    if int(sys.version[0]) < 3:
        python_version = 2
    else:
        python_version = 3
    logging.info("python_version = " + str(python_version))
    return python_version


def check_virtualenv_arg():
    global PIP_ACTIVATE, PIP_DEACTIVATE, venv_tmp_dir

    if PIP_ACTIVATE == "UNSPECIFIED":
        is_requirements_file = os.path.isfile("requirements.txt")

        if is_requirements_file != 1:
            logging.error("### Error Message ###")
            logging.error("Cannot find the virtualenv directory:" + PIP_ACTIVATE)
            logging.error("Also it cannot find 'requirements.txt' file and install pip package.")
            logging.error("Please check the '-a' option argument.")
            sys.exit(1)



        python_version = check_python_version()

        venv_path = os.path.join(CUR_PATH,venv_tmp_dir)

        if python_version == 2:
            create_venv_command = "virtualenv -p python " + venv_tmp_dir
        else:
            create_venv_command = "virtualenv -p python3 " + venv_tmp_dir

        if check_os() == "Windows":
            activate_command = ".\\" + os.path.join(venv_tmp_dir, "Scripts", "activate")
        else:
            activate_command = ". " + os.path.join(venv_path, "bin", "activate")

        PIP_ACTIVATE = activate_command

        install_command = "pip install -r requirements.txt"
        deactivate_command = "deactivate"
        PIP_DEACTIVATE = deactivate_command

        logging.info("You didn't enter the '-a' option.")
        logging.info("It makes virtualenv tmp dir(" + venv_path + ") to install pip package with requirements.txt.")

        if check_os() == "Windows":
            command_separator = "&"
        else:
            command_separator = ";"
        command_list = [create_venv_command, activate_command, install_command, deactivate_command]
        command = command_separator.join(command_list)
        command_ret = subprocess.call(command, shell=True)
        if command_ret != 0:
            logging.error("### Error Message ###")
            logging.error("This command(" + command + ") returns an error")
            logging.error("Please check if you installed virtualenv.")
            sys.exit(1)


def install_pip_package(package):
    pip_install_command = 'pip install ' + package
    ret = os.system(pip_install_command)

    if ret == 0:
        return True
    else:
        return False


def uninstall_pip_package(package):
    pip_uninstall_command = 'pip uninstall ' + package
    ret = os.system(pip_uninstall_command)

    if ret == 0:
        return True
    else:
        return False


def add_plugin_in_pom():
    global pom_backup

    is_append = False

    if os.path.isfile(manifest_array[2][1]) != 1:
        logging.error(manifest_array[2][1] + " is not existed in this directory.")
        sys.exit(1)

    shutil.move(manifest_array[2][1], pom_backup)

    license_maven_plugin = '<plugin>\
                                    <groupId>org.codehaus.mojo</groupId>\
                                    <artifactId>license-maven-plugin</artifactId>\
                                    <version>2.0.0</version>\
                                    <executions>\
                                        <execution>\
                                            <id>aggregate-download-licenses</id>\
                                            <goals>\
                                                <goal>aggregate-download-licenses</goal>\
                                            </goals>\
                                        </execution>\
                                    </executions>\
                                </plugin>'

    tmp_plugin = BeautifulSoup(license_maven_plugin, 'xml')

    license_maven_plugins = '<plugins>' + license_maven_plugin + '<plugins>'
    tmp_plugins = BeautifulSoup(license_maven_plugins, 'xml')

    with open(pom_backup, 'r', encoding='utf8') as f:
        f_xml = f.read()
        f_content = BeautifulSoup(f_xml, 'xml')

        build = f_content.find('build')
        if build is not None:
            plugins = build.find('plugins')
            if plugins is not None:
                plugins.append(tmp_plugin.plugin)
                is_append = True
            else:
                build.append(tmp_plugins.plugins)
                is_append = True

    if is_append:
        with open(manifest_array[2][1], "w", encoding='utf8') as f_w:
            f_w.write(f_content.prettify(formatter="minimal").encode().decode('utf-8'))

    return is_append


def clean_run_maven_plugin_output():
    global input_file_name

    if OUTPUT_CUSTOM_DIR != "":
        input_tmp = input_file_name.split('/')
        input_rest_tmp = ''
        for i in range(len(input_tmp)):
            if i > 0:
                input_rest_tmp = input_rest_tmp + '/' + input_tmp[i]
        input_file_name = str(OUTPUT_CUSTOM_DIR) + str(input_rest_tmp)

    directory_name = os.path.dirname(input_file_name)
    licenses_path = directory_name + '/licenses'
    if os.path.isdir(licenses_path) == 1:
        shutil.rmtree(licenses_path)
        os.remove(directory_name + '/licenses.xml')
        logging.info('remove temporary directory: ' + licenses_path)

    if len(os.listdir(directory_name)) == 0:
        shutil.rmtree(directory_name)

    shutil.move(pom_backup, manifest_array[2][1])


def run_maven_plugin():
    logging.info('run maven license scanning plugin with temporary pom.xml')
    command = "mvn license:aggregate-download-licenses"

    ret = subprocess.call(command, shell=True)

    if ret != 0:
        logging.error("### Error Message ###")
        logging.error("This command(" + command + ") returns an error.")

        clean_run_maven_plugin_output()
        sys.exit(1)


def open_input_file():
    global input_file_name

    if OUTPUT_CUSTOM_DIR != "":
        input_tmp = input_file_name.split('/')
        input_rest_tmp = ''
        for i in range(len(input_tmp)):
            if i > 0:
                input_rest_tmp = input_rest_tmp + '/' + input_tmp[i]
        input_file_name = str(OUTPUT_CUSTOM_DIR) + str(input_rest_tmp)

    if os.path.isfile(input_file_name) != 1:
        logging.error("### Error Message ###")
        logging.error(input_file_name + " doesn't exist in this directory.")

        if PACKAGE == "maven":
            global is_maven_first_try

            if is_maven_first_try:
                is_append = add_plugin_in_pom()
                is_maven_first_try = False

                if is_append:
                    run_maven_plugin()
                    return open_input_file()
                else:
                    clean_run_maven_plugin_output()

        logging.error("Please check the below thing first.")
        logging.error("  1.Did you run the OSS that scanning license of dependencies? If you didn't, then please refer the collab(http://collab.lge.com/main/x/MivvN) and try it.")
        logging.error("  2.Or if your project has the customized build output directory, then use '-o' option with your customized build output directory name")
        logging.error("     The default build output directory of maven is 'target', and which of gradle is 'build'.")
        logging.error("     For example, if your customized build output directory name is 'output', then run the below command.")
        logging.error("     >> $ python dependency_unified.py -o output")
        logging.error("  3.If you already checked 1st step and also it didn't work, then please request this issue through OSC CLM(http://clm.lge.com/issue/browse/OSC).")
        sys.exit(1)

    input_fp = open(input_file_name, 'r', encoding='utf8')

    return input_fp


def close_input_file(input_fp):
    input_fp.close()


def make_custom_json(tmp_custom_json):
    with open(tmp_custom_json, 'w', encoding='utf8') as custom:
        custom.write(
            "{\n\t\"name\": \"\",\n\t\"version\": \"\",\n\t\"licenses\": \"\",\n\t\"repository\": \"\",\n\t\"url\": \"\",\n\t\"copyright\": \"\",\n\t\"licenseText\": \"\"\n}\n".encode().decode(
                "utf-8"))


def start_license_checker():
    tmp_custom_json = "custom.json"
    tmp_file_name = "tmp_npm_license_output.json"
    flag_tmp_node_modules = False

    if os.path.isfile(tmp_custom_json) == 1:
        os.remove(tmp_custom_json)
    if os.path.isfile(tmp_file_name) == 1:
        os.remove(tmp_file_name)

    test_command = "license-checker > test.tmp"
    ret = os.system(test_command)
    os.remove("test.tmp")
    if ret != 0:
        logging.error("### Error Message ###")
        logging.error("Running license-checker returns error. Please check if the license-checker is installed.")
        logging.error(">>Command for installing license-checker (Root permission is required to run this command.)")
        logging.error("  sudo npm install -g license-checker")
        sys.exit(1)

    if os.path.isdir("node_modules") != 1:
        logging.info("node_modules directory is not existed.it executes 'npm install'.")
        flag_tmp_node_modules = True
        command = 'npm install'
        command_ret = subprocess.call(command, shell=True)
        if command_ret != 0:
            logging.error("### Error Message ###")
            logging.error("This command(" + command + ") returns an error")
            sys.exit(1)

    # customized json file for obtaining specific items with license-checker
    make_custom_json(tmp_custom_json)

    # license-checker option
    # --production : not prints devDependencies
    # --json : prints output file with json format
    # --out : output file name with path
    # --customPath : add the specified items to the usual ones
    command = 'license-checker --production --json --out ' + tmp_file_name + ' --customPath ' + tmp_custom_json
    command_ret = os.system(command)
    os.remove(tmp_custom_json)
    if flag_tmp_node_modules:
        shutil.rmtree('node_modules')

    return tmp_file_name


def start_pip_licenses():
    global PIP_ACTIVATE

    tmp_file_name = "tmp_pip_license_output.json"

    if PIP_ACTIVATE.startswith("source "):
        tmp_activate = PIP_ACTIVATE[7:]
        PIP_ACTIVATE = ". " + tmp_activate
    elif PIP_ACTIVATE.startswith("conda "):
        if check_os() == "Linux":
            tmp_activate = "eval \"$(conda shell.bash hook)\""
            PIP_ACTIVATE = tmp_activate + PIP_ACTIVATE

    if check_os() == 'Windows':
        command_separator = "&"
    else:
        command_separator = ";"
    activate_command = PIP_ACTIVATE
    install_pip_command = "pip install pip-licenses"
    pip_licenses_command = "pip-licenses --from=mixed --with-url --format=json --with-license-file > " + tmp_file_name
    uninstall_pip_command = "pip uninstall -y pip-licenses PTable"
    deactivate_command = PIP_DEACTIVATE

    command_list = [activate_command, install_pip_command, pip_licenses_command, uninstall_pip_command,
                    deactivate_command]
    command = command_separator.join(command_list)

    ret = subprocess.call(command, shell=True)

    if ret == 0:
        return tmp_file_name
    else:
        logging.error("### Error Message ###")
        logging.error("This command(" + command + ") returns an error.")
        sys.exit(1)


def check_os():
    # return value : Linux, Windows, Darwin(Mac OS)
    return platform.system()


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def check_license_scanner(os_name):
    global license_scanner_url, license_scanner_bin
    
    if os_name == 'Linux':
        license_scanner_url = license_scanner_url_linux
    elif os_name == 'Darwin':
        license_scanner_url = license_scanner_url_macos
    elif os_name == 'Windows':
        license_scanner_url = license_scanner_url_windows
    else:
        logging.info("Not supported OS to analyze license text with binary.")
        return

    license_scanner_bin = os.path.basename(license_scanner_url)


def check_to_exist_license_scanner():
    global license_scanner_bin, license_scanner_url

    fileobject = requests.get(license_scanner_url)
    if fileobject.status_code != 200:
        logging.error("### Error Message ###")
        logging.error("Downloading " + license_scanner_bin + " is failed.")
        return False
    else:
        with open(license_scanner_bin, 'wb') as f:
            f.write(fileobject.content)
    
    if os.path.isfile(license_scanner_bin) == 1:
        os.chmod(license_scanner_bin,0o755)
        return True
    else:
        return False


def check_and_run_license_scanner(file_dir, os_name):
    global license_scanner_first_flag, is_license_scanner, license_scanner_bin

    if license_scanner_first_flag:
        is_license_scanner = check_to_exist_license_scanner()
        license_scanner_first_flag = False

    if is_license_scanner:
        tmp_output_file_name = "tmp_license_scanner_output.txt"

        if file_dir == "UNKNOWN":
            license_name = ""
        else:
            if os_name == 'Linux':
                run_license_scanner = "./" + license_scanner_bin + " " + file_dir + " > " + tmp_output_file_name
            elif os_name == 'Darwin':
                run_license_scanner = "./" + license_scanner_bin + " identify " + file_dir + " > " + tmp_output_file_name
            elif os_name == 'Windows':
                run_license_scanner = license_scanner_bin + " identify " + file_dir + " > " + tmp_output_file_name
            else:
                run_license_scanner = ''

            if run_license_scanner is None:
                license_name = ""
                return license_name
            else:
                ret = os.system(run_license_scanner)
                if ret != 0:
                    return ""

            fp = open(tmp_output_file_name, "r", encoding='utf8')
            license_output = fp.read()
            fp.close()
            os.remove(tmp_output_file_name)

            if os_name == 'Linux':
                license_output_re = re.findall(r'.*contains license\(s\)\s(.*)', license_output)
            else:
                license_output_re = re.findall(r"License:\s{1}(\S*)\s{1}", license_output)

            if len(license_output_re) == 1:
                license_name = license_output_re[0]
                if license_name == "No_license_found":
                    license_name = ""
            else:
                license_name = ""

    else:
        license_name = ""

    return license_name


def check_multi_license(license_name):
    if isinstance(license_name, list):
        multi_license = 1
    else:
        multi_license = 0

    return multi_license


def parse_oss_name_version_in_filename(name):
    filename = name.rstrip('.jar')
    split_name = filename.rpartition('-')

    oss_name = split_name[0]
    oss_version = split_name[2]

    return oss_name, oss_version


def parse_oss_name_version_in_artifactid(name):
    artifact_comp = name.split(':')

    group_id = artifact_comp[0]
    artifact_id = artifact_comp[1]
    oss_version = artifact_comp[2]

    return group_id, artifact_id, oss_version


def version_refine(version):
    version_cmp = version.upper()

    if version_cmp.find(".RELEASE") != -1:
        version = version_cmp.rstrip(".RELEASE")
    elif version_cmp.find(".FINAL") != -1:
        version = version_cmp.rstrip(".FINAL")

    return version


#################################################
# Functions for generating OSS report xlsx file #
#################################################

def generate_oss_report():
    wb = Workbook()
    ws = wb.active

    ws.append(OSS_report_1st_row)
    ws.append(OSS_report_2nd_row)

    return wb


def save_oss_report(wb):
    wb.save(output_file_name)


def insert_oss_report(ws, data):
    # data format
    # [idx,package,oss_name,oss_version,license_name,dn_loc,homepage,copyright_text,license_text,exclude,comment]
    ws.append(data)


def load_oss_report_sheet():
    if os.path.isfile(output_file_name) != 1:
        logging.error("### Error Message ###")
        logging.error(output_file_name + " doesn't exist in this directory.")
        exit(1)

    wb = load_workbook(output_file_name, data_only=True)
    ws = wb.get_sheet_by_name('Sheet')

    return wb, ws


#################################################################################################################################
# Functions for parsing generated file from OSS that could scan license information about dependencies of each package manager.  #
#################################################################################################################################

def check_UNKNOWN(text):
    if text == ['UNKNOWN'] or text == 'UNKNOWN':
        text = ""

    return text


def parse_and_generate_output_pip(tmp_file_name):
    global license_scanner_bin

    os_name = check_os()
    check_license_scanner(os_name)

    try:
      with open(tmp_file_name, 'r', encoding='utf-8') as json_file:
        json_data = json.load(json_file)

        wb = generate_oss_report()

        idx = 1
        for d in json_data:
            oss_init_name = d['Name']
            oss_name = "pypi:" + oss_init_name
            license_name = check_UNKNOWN(d['License'])
            homepage = check_UNKNOWN(d['URL'])
            oss_version = d['Version']
            dn_loc = dn_url + oss_init_name + "/" + oss_version

            license_file_dir = d['LicenseFile']
            license_name_with_license_scanner = check_and_run_license_scanner(license_file_dir, os_name)

            if license_name_with_license_scanner != "":
                license_name = license_name_with_license_scanner

            insert_oss_report(wb.active,
                              [str(idx), 'pip', oss_name, oss_version, license_name, dn_loc, homepage, '', '', '', ''])
            idx += 1

        save_oss_report(wb)

    except Exception as ex:
        logging.info("Error:"+ str(ex))

    if os_name == 'Linux':
        if is_license_scanner:
            os.remove(license_scanner_bin)

    if os.path.isdir(venv_tmp_dir):
        shutil.rmtree(venv_tmp_dir)
        logging.info("remove tmp directory: " + venv_tmp_dir)


def parse_and_generate_output_npm(tmp_file_name):
    with open(tmp_file_name, 'r', encoding='utf8') as json_file:
        json_data = json.load(json_file)

        wb = generate_oss_report()

        keys = [key for key in json_data]

        idx = 1
        for i in range(0, len(keys)):
            d = json_data.get(keys[i - 1])
            oss_init_name = d['name']
            oss_name = "npm:" + oss_init_name

            if d['licenses']:
                license_name = d['licenses']
            else:
                license_name = ''

            if d['repository']:
                dn_loc = d['repository']
            else:
                dn_loc = dn_url + oss_init_name

            homepage = dn_url + oss_init_name
            oss_version = d['version']

            if d['copyright']:
                copyright_text = d['copyright']
            else:
                copyright_text = ''

            multi_license = check_multi_license(license_name)

            if multi_license == 1:
                for l_idx in range(0, len(license_name)):
                    license_name = license_name[l_idx].replace(",", "")

                    insert_oss_report(wb.active,
                                      [str(idx), 'package.json', oss_name, oss_version, license_name, dn_loc,
                                       homepage, copyright_text, '', '', ''])
                    idx += 1
            else:
                license_name = license_name.replace(",", "")

                insert_oss_report(wb.active,
                                  [str(idx), 'package.json', oss_name, oss_version, license_name, dn_loc, homepage,
                                   copyright_text, '', '', ''])
                idx += 1

        save_oss_report(wb)


def parse_and_generate_output_maven(input_fp):
    tree = parse(input_fp)

    root = tree.getroot()
    dependencies = root.find("dependencies")

    wb = generate_oss_report()
    idx = 1

    for d in dependencies.iter("dependency"):
        groupid = d.findtext("groupId")
        artifactid = d.findtext("artifactId")
        version = d.findtext("version")
        oss_version = version_refine(version)

        oss_name = groupid + ":" + artifactid

        license_num = 1
        licenses = d.find("licenses")
        if len(licenses):
            for key_license in licenses.iter("license"):
                license_name = key_license.findtext("name")
                dn_loc = dn_url + groupid + "/" + artifactid + "/" + version
                homepage = dn_url + groupid + "/" + artifactid

                license_name = license_name.replace(",", "")

                insert_oss_report(wb.active,
                                  [str(idx), 'pom.xml', oss_name, oss_version, license_name, dn_loc, homepage, '', '',
                                   '', ''])

                license_num += 1
                idx += 1
        else:
            # Case that doesn't include License tag value.
            license_name = ''
            dn_loc = dn_url + groupid + "/" + artifactid + "/" + version
            homepage = dn_url + groupid + "/" + artifactid

            insert_oss_report(wb.active,
                              [str(idx), 'pom.xml', oss_name, oss_version, license_name, dn_loc, homepage, '', '', '',
                               ''])

            idx += 1

    save_oss_report(wb)


def parse_and_generate_output_gradle(input_fp):
    json_data = json.load(input_fp)

    wb = generate_oss_report()
    idx = 1

    for d in json_data['dependencies']:

        used_filename = "false"
        group_id = ""
        artifact_id = ""

        name = d['name']
        filename = d['file']

        if name != filename:
            group_id, artifact_id, oss_ini_version = parse_oss_name_version_in_artifactid(name)
            oss_name = group_id + ":" + artifact_id
        else:
            oss_name, oss_ini_version = parse_oss_name_version_in_filename(filename)
            used_filename = "true"

        oss_version = version_refine(oss_ini_version)

        for licenses in d['licenses']:
            license_name = licenses['name']

            license_name = license_name.replace(",", "")

            if used_filename == "true" or group_id == "":
                dn_loc = 'Unknown'
                homepage = ''

            else:
                dn_loc = dn_url + group_id + "/" + artifact_id + "/" + oss_ini_version
                homepage = dn_url + group_id + "/" + artifact_id

            insert_oss_report(wb.active,
                              [str(idx), 'build.gradle', oss_name, oss_version, license_name, dn_loc, homepage, '', '',
                               '', ''])
            idx += 1

    save_oss_report(wb)


def preprocess_pub_result(input_file):
    matched_json = re.findall(r'final ossLicenses = <String, dynamic>({.*});', input_file.read())

    if matched_json[0] is not None:
        return matched_json[0]
    else:
        logging.error("### Error Message ###")
        logging.error("Cannot parse the result json from pub input file.")
        exit(1)


def parse_and_generate_output_pub(tmp_file_name):
    global license_scanner_bin, tmp_license_txt_file_name

    json_txt = preprocess_pub_result(tmp_file_name)
    json_data = json.loads(json_txt)

    wb = generate_oss_report()

    os_name = check_os()
    check_license_scanner(os_name)

    idx = 1
    for key in json_data:
        oss_name = json_data[key]['name']
        oss_version = json_data[key]['version']
        homepage = json_data[key]['homepage']
        dn_loc = homepage
        license_txt = json_data[key]['license']

        tmp_license_txt = open(tmp_license_txt_file_name, 'w', encoding='utf-8')
        tmp_license_txt.write(license_txt)
        # tmp_license_txt.write(license_txt.encode().decode('utf-8'))
        tmp_license_txt.close()

        license_name_with_license_scanner = check_and_run_license_scanner(tmp_license_txt_file_name, os_name)

        if license_name_with_license_scanner != "":
            license_name = license_name_with_license_scanner
        else:
            license_name = ''

        insert_oss_report(wb.active,
                          [str(idx), 'pub', oss_name, oss_version, license_name, dn_loc, homepage, '', '', '', ''])
        idx += 1

    save_oss_report(wb)

    if os_name != 'Windows':
        if is_license_scanner:
            os.remove(license_scanner_bin)

    os.remove(tmp_license_txt_file_name)


###########################################
# Main functions for each package manager  #
###########################################

def main_pip():
    # It needs the virtualenv path that pip packages are installed.
    check_virtualenv_arg()

    # Run the command 'pip-licenses with option'.
    tmp_file_name = start_pip_licenses()

    # Make output file for OSS report using temporary output file for pip-licenses.
    parse_and_generate_output_pip(tmp_file_name)

    # Remove temporary output file.
    if os.path.isfile(tmp_file_name):
        os.remove(tmp_file_name)


def main_npm():
    # Install the license-checker (npm package) with global option.
    # os.system("npm install -g license-checker")

    # Run the command 'license-checker' with option'.
    tmp_file_name = start_license_checker()

    # Make output file for OSS report using temporary output file for license-checker.
    parse_and_generate_output_npm(tmp_file_name)

    # Remove temporary output file.
    os.remove(tmp_file_name)


def main_maven():
    # Before running this script, first you should add the maven-license-plugin in pom.xml and run it.

    # open license.xml
    input_fp = open_input_file()

    # Make output file for OSS report using temporary output file for maven-license-plugin.
    parse_and_generate_output_maven(input_fp)

    # close licenses.xml
    close_input_file(input_fp)

    if not is_maven_first_try:
        clean_run_maven_plugin_output()


def main_gradle():
    # Before running this script, first you should add the com.github.hierynomus.license in build.gradle and run it.

    # open dependency-license.json
    input_fp = open_input_file()

    # Make output file for OSS report using temporary output file for License Gradle Plugin.
    parse_and_generate_output_gradle(input_fp)

    # close dependency-license.json
    close_input_file(input_fp)


def main_pub():
    input_fp = open_input_file()

    parse_and_generate_output_pub(input_fp)

    close_input_file(input_fp)


def main():
    # Global variables
    global PACKAGE, output_file_name, input_file_name, CUR_PATH, OUTPUT_RESULT_DIR, MANUAL_DETECT, OUTPUT_CUSTOM_DIR, dn_url, PIP_ACTIVATE, PIP_DEACTIVATE
    global license_scanner_first_flag, is_license_scanner, license_scanner_url, license_scanner_bin, venv_tmp_dir, pom_backup, is_maven_first_try, tmp_license_txt_file_name

    # Init logging
    logging.basicConfig(level=logging.INFO, format='%(message)s')

    # Configure global variables according to package manager.
    try:
        configure_package()
    except HelpStop:
        sys.exit(0)
    except:
        logging.error("### Error Message ###")
        help_print()
        sys.exit(1)

    if PACKAGE == "pip":
        dn_url = "https://pypi.org/project/"
        output_file_name = "pip_dependency_output.xlsx"
        venv_tmp_dir = "venv_osc_dep_tmp"
        license_scanner_first_flag = True
        is_license_scanner = False

    elif PACKAGE == "npm":
        dn_url = "https://www.npmjs.com/package/"
        output_file_name = "npm_dependency_output.xlsx"

    elif PACKAGE == "maven":
        dn_url = "https://mvnrepository.com/artifact/"
        input_file_name = "target/generated-resources/licenses.xml"
        output_file_name = "maven_dependency_output.xlsx"
        pom_backup = "pom.xml_backup"
        is_maven_first_try = True

    elif PACKAGE == "gradle":
        dn_url = "https://mvnrepository.com/artifact/"
        input_file_name = "build/reports/license/dependency-license.json"
        output_file_name = "gradle_dependency_output.xlsx"

    elif PACKAGE == "pub":
        input_file_name = "lib/oss_licenses.dart"
        output_file_name = "pub_dependency_output.xlsx"
        license_scanner_first_flag = True
        is_license_scanner = False
        tmp_license_txt_file_name = "tmp_license.txt"

    else:
        logging.error("### Error Message ###")
        logging.error("You enter the wrong first argument.")
        logging.error("Please enter the package manager into (pip, npm, maven, gradle)")
        sys.exit(1)

    if PACKAGE == "pip":
        main_pip()
    elif PACKAGE == "npm":
        main_npm()
    elif PACKAGE == "maven":
        main_maven()
    elif PACKAGE == "gradle":
        main_gradle()
    elif PACKAGE == "pub":
        main_pub()

    logging.info("### FINISH!! ###")

    if os.path.isfile(output_file_name):
        shutil.move(output_file_name, OUTPUT_RESULT_DIR + "/" + output_file_name)
        logging.info("Generated " + output_file_name + " in " + OUTPUT_RESULT_DIR + "!")
    else:
        logging.info("No file was created due to an error.")

    sys.exit(0)


if __name__ == '__main__':
    main()
